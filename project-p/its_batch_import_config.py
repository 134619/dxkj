# -*- coding: utf-8 -*-

import importlib
import json
import os
import re
import subprocess
from datetime import datetime

import pandas as pd
from BaseConfig import get_config
from DbHelper import DbHelper
from its_config import DATA
from libenhance import Request, Response, SysHandler
from openpyxl import load_workbook
from ToolsMethods import ResetResponse, ResponseData, ResponseStatusCode
from UploadFile import FileToSys

# res = Response()
# req = Request()


# 批次导入的配置
BATCHCONFIG = [
        {
            "import_type_func": "purchase_receipt",
            "import_type_script": "itf_purchase_receipt",
            "import_type_description": "采购收货",
            "batch_import_type": "safe_purchase_receipt",
            "safe_path" : "./data/Vendor/Import/",
            "unique_key" : "4cb30693",
            "type_code": "K02"
        },
        {
            "import_type_func": "ito_post",
            "import_type_script": "itf_ito_post",
            "safe_path" : "./data/Vendor/Import/",
            "unique_key" : "8b44a78e",
            "import_type_description": "公司间调拨过账",
            "batch_import_type": "safe_inter_company_transfer",
            "type_code": "S05"
        },
        {
            "import_type_func": "delivery_posting",
            "import_type_script": "itf_delivery_posting",
            "unique_key" : "340fe4d1",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "发货过账",
            "batch_import_type": "safe_delivery_posting",
            "type_code": "S07"
        },
        {
            "import_type_func": "emes_push_mh_offline_data",
            "import_type_script": "itf_emes_offline_data",
            "unique_key" : "340fe4d1",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "下线绑批",
            "batch_import_type": "safe_batch_bind",
            "type_code": "P03"
        },
        {
            "import_type_func": "cof_cutoff_downline",
            "import_type_script": "itf_cof_cutoff_downline",
            "unique_key" : "f7d0bb09",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "DS切割标记",
            "batch_import_type": "safe_cut_mark",
            "type_code": "K05"
        },
        {
            "import_type_func": "process_ds_loss",
            "import_type_script": "itf_ds_loss",
            "safe_path" : "./data/Vendor/Import/",
            "unique_key" : "28ca9a6f",
            "import_type_description": "DS损耗过账",
            "batch_import_type": "safe_ds_loss",
            "type_code": "K07"
        },
        {
            "import_type_func": "subcontract_receipt",
            "import_type_script": "itf_subcontract_receipt",
            "unique_key" : "652e82c4",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "外协收货扣料",
            "batch_import_type": "safe_subcontract_receipt",
            "type_code": "K09"
        },
        {
            "import_type_func": "subcontract_receipt_reverse",
            "import_type_script": "itf_subcontract_receipt_reverse",
            "unique_key" : "2bdc500a",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "外协收货扣料冲销",
            "batch_import_type": "safe_subcontract_receipt_reverse",
            "type_code": "K11"
        },
        {
            "import_type_func": "save_emes_wanjia_goods_transfer_data",
            "import_type_script": "itf_emes_wanjia_goods_transfer_data",
            "unique_key" : "a4aaf470",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "库存调拨过账（万佳）",
            "batch_import_type": "safe_stock_transfer_wanjia",
            "type_code": "K12"
        },
        {
            "import_type_func": "stock_batch_import_osat",
            "import_type_script": "itf_stock_batch_import_osat",
            "unique_key" : "b2c216e9",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "库存调拨批导（OSAT）",
            "batch_import_type": "safe_stock_batch_import_osat",
            "type_code": "K14"
        },
        {
            "import_type_func": "internal_receive",
            "import_type_script": "itf_internal_receive",
            "unique_key" : "350a2b09",
            "safe_path" : "./data/Vendor/Import/",
            "import_type_description": "内部领用",
            "batch_import_type": "safe_internal_receive",
            "type_code": "K15"
        }
    ]




def get_safe_path_by_batch_type(batch_import_type):
    for item in BATCHCONFIG:
        if item["batch_import_type"] == batch_import_type:
            return item["safe_path"]
    # 没找到返回 None
    return None


# def get_map_its(processed_data):
#     # 构建快速查找映射表 { excel_header: its_header }
#     header_map = {item["excel_header"]: item["its_header"] for item in DATA}

#     # 生成新的 processed_data（key 替换为 its_header）
#     new_processed_data = []
#     for item in processed_data:
#         new_item = {}
#         for old_key, value in item.items():
#             new_key = header_map.get(old_key, old_key)  # 找不到则保留原key
#             new_item[new_key] = value
#         new_processed_data.append(new_item)
#     mapping_list = [{"field": "type", "description": "执行结果", "type": "varchar", "length": 20, "decimal": 0},{"field": "msg", "description": "消息文本", "type": "varchar", "length": 200, "decimal": 0}]
#     for item in DATA:
#         excel_header = item["excel_header"]
#         its_header = item["its_header"]
#         # 判断：当前 excel_header 是否存在于 processed_data 的 key 中
#         key_exists = any(excel_header in data_item for data_item in processed_data)
#         if key_exists:
#             mapping_list.append({
#                 "description": excel_header,
#                 "field": its_header
#             })
#     return new_processed_data,mapping_list

def get_map_its(processed_data):
    # 构建快速查找映射表 { excel_header: its_header }
    header_map = {item["excel_header"]: item["its_header"] for item in DATA}
    
    # 反向映射表，用于根据 its_header 反查 excel_header（用于 mapping_list 的 description）
    reverse_map = {item["its_header"]: item["excel_header"] for item in DATA}

    # 生成新的 processed_data（key 替换为 its_header）
    new_processed_data = []
    for item in processed_data:
        new_item = {}
        for old_key, value in item.items():
            new_key = header_map.get(old_key, old_key)
            new_item[new_key] = value
        new_processed_data.append(new_item)

    # 初始化 mapping_list（固定部分）
    mapping_list = [
        {"field": "type", "description": "执行结果", "type": "varchar", "length": 20, "decimal": 0},
        {"field": "msg", "description": "消息文本", "type": "varchar", "length": 200, "decimal": 0}
    ]

    # 用于去重的 set，记录已经添加过的 its_header
    added_fields = set()

    # 按照 processed_data 第一行数据的原始 key 顺序遍历（即 Excel 导入的字段顺序）
    if processed_data and len(processed_data) > 0:
        first_item = processed_data[0]
        for excel_header in first_item.keys():
            # 查找对应的 its_header
            its_header = header_map.get(excel_header)
            if its_header is None:
                continue  # 不在映射表中的字段跳过
            # 去重判断
            if its_header in added_fields:
                continue
            
            # 添加到 mapping_list
            mapping_list.append({
                "description": excel_header,
                "field": its_header
            })
            added_fields.add(its_header)

    return new_processed_data, mapping_list

# @ResetResponse()
# def GetSupplyType():
#     # 下拉接口
#     data = BATCHCONFIG
#     print(f"------------{data}")
#     # res.set_body(json.dumps({"code": 200, "msg": "ok", "data": data}))
#     # res.commit(True)
#     result={
#         "test":data
#     }
#     return  ResponseData(ResponseStatusCode.Success, "成功",data)
res = Response()
req = Request()

def GetSupplyType():
    # 下拉接口
    data = BATCHCONFIG
    res.set_body(json.dumps({"code": 200, "msg": "ok", "data": data}))
    res.commit(True)


def UploadFile():
    import_type_func = req.multipartformdata("import_type_func")[0]["content"]  # 方法名称
    import_type_script = req.multipartformdata("import_type_script")[0]["content"]  # 文件名称
    batch_import_type = req.multipartformdata("batch_import_type")[0]["content"]
    target_path = get_safe_path_by_batch_type(batch_import_type)
    file_data_list = req.multipartformdata("file")
    processed_data = []
    code, msg, file_name, data, display = 200, "上传成功", [], [], []
    file_name_list =  []
    user_id = req.header("user_id") or sys.get_env("user_id") or 0
    print("--user_id--",user_id)
    for file_data in file_data_list:
        if "." not in file_data["filename"]:
            res.set_body(
                json.dumps(
                    {"code": 400, "msg": f"文件名无后缀, 错误名称为{file_data['filename']}, 请检查", "data": []}))
            res.commit(True)
            return
        if file_data["filename"].count(".") > 1:
            res.set_body(json.dumps(
                {"code": 400, "msg": f"文件名中不可含有多个符号点, 错误名称为{file_data['filename']}", "data": []}))
            res.commit(True)
            return
        file_name, file_type = file_data["filename"].split(".")
        if file_type.lower() not in ["xlsx", "zip", "mhtml"]:
            res.set_body(json.dumps(
                {"code": 400, "msg": f"请上传后缀为xlsx/zip的文件, 错误名称为{file_data['filename']}", "data": []}))
            res.commit(True)
            return
        if file_name[-14:].startswith("20") and file_name[-14:].isdigit():
            file_name = file_name + "." + file_type
        else:
            file_name = file_name + datetime.now().strftime("%Y%m%d%H%M%S") + "." + file_type
        file_name, file_type = file_data["filename"].split(".")
        if file_name[-14:].startswith("20") and file_name[-14:].isdigit():
            file_name = file_name + "." + file_type
        else:
            file_name = file_name + datetime.now().strftime("%Y%m%d%H%M%S") + "." + file_type
        file_path = target_path + file_name
        print("上传文件的文件路径",file_path)
        os.system('mkdir -p {}'.format(target_path))
        print("移动的命令",'mkdir -p {}'.format(target_path))
        os.system('mv "{}" "{}"'.format(file_data["content"], file_path))
        print("mv的命令",'mv "{}" "{}"'.format(file_data["content"], file_path))

        if not os.path.exists(file_path):
            res.set_body(
                json.dumps({"code": 400, "msg": "目标文件不存在, 请检查", "data": []}))
            res.commit(True)
            return
        code = 200
        msg = '上载成功'
        display = []
        ########################## 读取数据#############
        wb = load_workbook(file_path)
        active_sheet_name = wb.active.title  # 获取激活表名
        wb.close()
        error_index_dic = {}
        error_index_dic['sheet_names'] = [active_sheet_name]
        columns_error = {}
        # 读取激活Sheet
        df = pd.read_excel(file_path, sheet_name=active_sheet_name, engine='openpyxl', dtype=str,
                           usecols=lambda x: x != "").dropna(how='all', axis=0)
        # # 删除第一列
        df.fillna("", inplace=True)
        columns_list = df.columns.tolist()
        # 列记录
        for columns in columns_list:
            if active_sheet_name not in columns_error:
                columns_error[active_sheet_name] = []
            columns_error[active_sheet_name].append(columns)
        # 行数据转字典
        for item in df.values:
            data_dict = dict(zip(columns_list, item))
            processed_data.append(data_dict)

        # file_name_list =  []
        file_name_list.append(file_name)

    ###########################
    new_processed_data,display = get_map_its(processed_data)
    print("全部参数",new_processed_data)
    # try:
    #     print(3333333333333,import_type_script,import_type_func)
    #     module = importlib.import_module(import_type_script)
    #     # # 从模块里取函数
    #     func = getattr(module, import_type_func, None)
    #     if func:
    #         status_code, msg, data ,display = func(processed_data,user_id)
    #     if status_code == "S":
    #         status_code = 200
    #     else:
    #         status_code = 400
    # except Exception as e:
    #     print(1111,"调用方法不存在",e)
    #     # 调用方法不存在
    #     res.set_body(json.dumps({"code": 500, "msg": f"调用方法不存在: {str(e)}", "data": []}))
    #     res.commit(True)
    #     # is_can_import = False
    #     return
    ########################################################################################################################
    if code == 200:
        status_code = "S"
        msg_end = "上传成功"
    else:
        status_code = "E"
        msg_end = "上传失败"
    for item in new_processed_data:
        item["type"] = status_code
        item["msg"] = msg_end
    print(json.dumps({
        "code": code,
        "msg": msg,
        "file_name": file_name_list,
        "data": new_processed_data,
        "display": display,
        # "is_can_import": is_can_import,
        # "error_sheet": error_sheet
    }))
    if code != 200:
        os.system(f'rm -f "{file_path}"')
    res.set_body(json.dumps({
        "code": code,
        "msg": msg,
        "file_name": file_name_list,
        "data": new_processed_data,
        "display": display,
    }))
    res.commit(True)
    return






def ImportFileToDB():
    # 导入保存
    db = DbHelper()
    user_id = req.header("user_id")
    try:
        body = json.loads(req.body())
    except Exception as e:
        res.set_body(json.dumps({"code": 400, "msg": "请求参数格式错误", "data": []}))
        res.commit(True)
        return
    # 安全获取参数
    file_name_list = body.get("file_name", "")
    batch_import_type = body.get("batch_import_type", "")
    import_type_func = body.get("import_type_func", "")
    import_type_script = body.get("import_type_script", "")
    processed_data = []
    for file_name in file_name_list:
        if not all([file_name, batch_import_type]):
            res.set_body(json.dumps({"code": 400, "msg": "file_name、batch_import_type 不能为空", "data": []}))
            res.commit(True)
            return

        # 切割文件名，防止无后缀报错
        if "." not in file_name:
            res.set_body(json.dumps({"code": 400, "msg": "文件名必须包含后缀（如.xlsx）", "data": []}))
            res.commit(True)
            return
        msg = "保存成功"    
        file_name_only, file_type = file_name.rsplit(".", 1)  # 安全切割

        # # 查询是否已上传
        # print("查询是否上传",f"""SELECT *  FROM`t_general_import_log`WHERE`batch_import_type` = '{batch_import_type}'AND `file_name` = '{file_name_only}'AND `file_type` = '{file_type}'""")
        # data = db.query_sql(f"""
        #     SELECT
        #         *
        #     FROM
        #         `t_general_import_log`
        #     WHERE
        #         `batch_import_type` = '{batch_import_type}'
        #     AND `file_name` = '{file_name_only}'
        #     AND `file_type` = '{file_type}'
        # """)

        # if data:
        #     res.set_body(json.dumps({"code": 400, "msg": "该文件已上传,如需重新上传请避免重名", "data": []}))
        #     res.commit(True)
        #     return

        target_path = get_safe_path_by_batch_type(batch_import_type)

        file_path = target_path + file_name
        file_name_list =  []
    ########################## 读取数据#############
        wb = load_workbook(file_path)
        active_sheet_name = wb.active.title  # 获取激活表名
        wb.close()
        error_index_dic = {}
        error_index_dic['sheet_names'] = [active_sheet_name]
        columns_error = {}
        df = pd.read_excel(file_path, sheet_name=active_sheet_name, engine='openpyxl', dtype=str,
                           usecols=lambda x: x != "").dropna(how='all', axis=0)
        df.fillna("", inplace=True)
        columns_list = df.columns.tolist()
        for columns in columns_list:
            if active_sheet_name not in columns_error:
                columns_error[active_sheet_name] = []
            columns_error[active_sheet_name].append(columns)
        for item in df.values:
            data_dict = dict(zip(columns_list, item))
            processed_data.append(data_dict)
        print("全部参数",processed_data)
    ###########################
        # 映射关系
        new_processed_data,display = get_map_its(processed_data)
        # 初始化返回数据
        status_code = 200
        result_data = []  # 统一返回数据变量
        try:
            # 动态调用导入函数
            module = importlib.import_module(import_type_script)
            func = getattr(module, import_type_func, None)
            if func:
                # 调用处理函数
                print("new_processed_data:",new_processed_data)
                result = func(new_processed_data,user_id)
                print("result:",result)
                status_code = result.get("type","E")
                msg = result.get("message","未返回msg")
            else:
                status_code = 500
                msg = f"未找到函数：{import_type_func}"

        except Exception as e:
            print("导入函数执行异常：", e)
            status_code = 500
            msg = f"导入失败调用函数报错：{str(e)}"
        print(file_path,'======')

        if not os.path.isfile(file_path):
            return 500, '{}文件路径不存在'.format(file_path)
        print("status_code",status_code)

        for item in new_processed_data:
            item["type"] = status_code
            item["msg"] = msg   

        if status_code == "S":
            status_code = 200
            msg_end = "保存成功"

        else:
            status_code = 500
            msg_end = "保存失败"

             
        if status_code == 200:
            command = """md5sum "{}" | cut -d' ' -f1""".format(file_path)
            result = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                                    universal_newlines=True)
            file_md = result.stdout
            db = DbHelper()
            save_sql = """INSERT INTO `t_general_import_log` (`batch_import_type`, `file_type`, `file_name`, `md5`, `file_status`, `create_id`, `update_id`) 
            VALUES ('{}', '{}', '{}', '{}', '存在', '{}', '{}')""".format(batch_import_type, file_type, file_name_only, file_md,
                                                                            user_id, user_id)
            print("导入日志的sql",save_sql)
            status = db.exec_sql(save_sql)
        file_name_list.append(file_name)
        
    res.set_body(json.dumps({
        "code": status_code,
        "msg": msg_end,
        "data": new_processed_data,  #
        "file_name": file_name_list,
        "display": display
    }, ensure_ascii=False))
    res.commit(True)



def DeleteFile():
    # 上传失败 删除文件
    # res = Response()
    # req = Request()

    body = json.loads(req.body())
    batch_import_type = body["batch_import_type"]
    file_name = body["file_name"]
    
    if not path:
        res.set_body(json.dumps({"code": 500, "msg": '未查询到{}批导类型对应的路径，请检查'.format(batch_import_type)}))
        res.commit(True)
        return
    file_path = os.path.join(path, file_name)

    os.system('rm -rf "{}"'.format(file_path))
    res.set_body(json.dumps({"code": 200, "msg": "删除成功!"}))
    res.commit(True)